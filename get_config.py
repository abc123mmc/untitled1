import re
from openpyxl  import load_workbook

class GetCofig():
    @staticmethod
    def get_layout(file):
        """从layout文件中读取指定页的所有字段及位置,参数[界面配置文件路径]"""
        wb = load_workbook(file)
        wb._active_sheet_index=0
        ws = wb.active
        row = ws[1]
        return {value.value:index+1 for index,value in enumerate(row)}

    @staticmethod
    def read_configure(file):
        """获取配置，返回各配置文件配置"""
        wb=load_workbook(file)
        ws=wb['人工处理']
        rgcl=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
        rgcl=[i for i in rgcl if i.count(None)!=len(i)]
        rgcl=[{i1:i2 for i1,i2 in zip(rgcl[0],i) if i2 and i1 !='备注'} for i in rgcl[1:]]
        ws=wb['改快递']
        gkd=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
        gkd=[i for i in gkd if i.count(None)!=len(i)]
        gkd=[{i1:i2 for i1,i2 in zip(gkd[0],i) if i2 and i1 !='备注'} for i in gkd[1:]]
        ws=wb['改赠品']
        gzp=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
        gzp=[i for i in gzp if i.count(None)!=len(i)]
        gzp=[{i1:i2 for i1,i2 in zip(gzp[0],i) if i2 and i1 !='备注'} for i in gzp[1:]]
        ws=wb['当前使用快递']
        k=[[i.value for i in ws[i+1]] for i in range(ws.max_row)][1:]
        k=[i for i in k if i.count(None)!=len(i)]
        k={i[0]:i[1] for i in k}
        wb.close()
        return rgcl,gkd,gzp,k

if __name__ == '__main__':
    pass
    import setting
    excel1 = r'自动审核执行文件/快递配置.xlsx'  # 快递配置
    z=GetCofig()
    print(setting.ddclzx)
    a=GetCofig.get_layout(setting.ddclzx)
    b=GetCofig.get_layout(setting.spxx)
    excel1=GetCofig.read_configure(excel1)
    pass


