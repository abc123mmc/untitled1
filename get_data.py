# -*- coding: utf-8 -*-
import re,time,xlrd
import pyautogui#模拟鼠标键盘操作
import pyperclip#剪切板操作
import traceback#用于错误处理
from openpyxl  import load_workbook

#******************************************取数据***********************************************
def diZhiChuLi(beizhu):
    '''地址处理:传入存在改地址的备注，返回address'''
    beizhu=beizhu.replace('\n','')#去除备注中的换行
    dz01=re.findall('改地址[:：](.*?)┋',beizhu)
    d=''
    if dz01:
        d=''.join(e for e in dz01[-1] if e.isalnum() or e in ':：;；._-')#去除备注里的符号
        c='(?:联系方式|电话|手机).{,5}?(?=1\d{10})|(?:所在地区|详细地址|收货人|姓名|地址)[:：]'
        d=re.sub(c,'',d)
    if d:d=re.split('[;；]',d)
    if len(d)>3:return False,False
    dh=[]
    xm=[]
    dz=[]
    for i in d:
        cd=len(i)
        sz=len(re.findall('\d',i))
        sj=re.findall('省|市|区|县|乡|镇|村|路|号|街|道|',i)
        try:
            while 1:sj.remove('')
        except:pass
        sj=len(sj)
        if sz>9:
            if cd-sz<6:dh+=[i]
            else:dz+=[i]
        else:
            if sj>2:dz+=[i]
            elif sj==0:
                if cd<8:xm+=[i]
                else:dz+=[i]
            else:
                if cd>10:dz+=[i]
                else:xm+=[i]
    for i in [dh,dz,xm]:
        try:
            while 1:i.remove('')
        except:pass
    if len(dh)==0:dh=''
    elif len(dh)==1:dh=dh[0]
    else:
        dh1=''
        for i in dh:
            if not dh1:
                num1=len(i)
                dh1=i
            elif num1>len(i):
                num1=len(i)
                dh1=i
        dh=dh1
    if len(xm)==0:xm=''
    elif len(xm)==1:xm=xm[0]
    else:
        xm1=''
        for i in xm:
            if not xm1:
                num1=len(i)
                xm1=i
            elif num1>len(i):
                num1=len(i)
                xm1=i
        xm=xm1
    if len(dz)==0:dz=''
    elif len(dz)==1:dz=dz[0]
    else:
        dz1=''
        for i in dz:
            if not dz1:
                num1=len(i)
                dz1=i
            elif num1<len(i):
                num1=len(i)
                dz1=i
        dz=dz1
    sqx=''
    if dz:
        t1= xlrd.open_workbook(r'地区.xlsx').sheets()[2]
        li=[]
        dz1=''.join(re.findall('.*(?<=[区县市旗岛域辖镇乡台阁仔])',dz))
        if not dz1:dz1=dz
        for i in range(t1.nrows):#循环行
            h=t1.row_values(i)
            zhi=0;zhi2=0
            for i1 in h[3:6]:#分析当前行的456格
                if i1 in dz1:zhi2+=5
                elif i1[:2] in dz1:zhi2+=3
                for i2 in range(len(i1)):#i1地址单元格
                    if i2<2 and i1[i2] in dz1:zhi+=2
                    elif i1[i2] in dz1:zhi+=1
            if zhi2<6:continue#权重不足时跳出当次循环
            zhi=zhi+zhi2
            if zhi>8:li.append([i,zhi])#i行，zhi权重
        zhi=[i[1] for i in li]#值的列表
        zhi1=sorted(zhi)[-2:]
        if not zhi1:print('改地址备注权重不足以判断地区：备注内容非地址',dz);return False,False
        elif len(zhi1)==1 or zhi1[0]!=zhi1[1]:
            if zhi1[-1]>10:sqx=t1.row_values(li[zhi.index(zhi1[-1])][0])[:5]
            else:print('改地址备注权重不足以判断地区：备注内容不全',dz);return False,False
        else:
            if zhi1[0]>10:
                g5=t1.row_values(li[zhi.index(zhi1[-1])][0])[4]
                #sqx=[t1.row_values(i)[:3] for i in range(t1.nrows) if (t1.row_values(i)[4]==g5 and '其它区' in t1.row_values(i)[5])][0]
                try:sqx=[t1.row_values(i)[:5] for i in range(t1.nrows)
                         if (t1.row_values(i)[4]==g5 and '其它区' in t1.row_values(i)[5])][0]
                except:print('其他区不存在1111111111111111',dz);return False,False
            else:print('修改地址备注异常222222222222',dz);return False,False
    shengshi=[]
    sqx1=[]
    if sqx:
        sqx1= [int(x) for x in sqx[:3]]
        shengshi= [x for x in sqx[3:]]
    address=(xm,dh,dz,sqx1)#姓名、电话、地址、地区编码的列表,省份城市列表
    for i in address:
        if i:return address,shengshi
    #print('电话',dh);print('姓名',xm);print('地址',dz);print('编码',sqx);print('省份城市',shengshi)
    return False,False

def xuanzeshijian():#时间段
    t1,li=time.time(),[1,4,8,12,24,72]
    lli=[]
    for i in range(len(li)-1):
        t2=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(t1-3600*li[i+1]))
        t3=time.strftime('%Y-%m-%d %H:%M:%S',time.localtime(t1-3600*li[i]))
        lli.append([t2,t3])
    return lli

def ddsj():#订单数据：去取当前查询到的当前页所有订单的数据并返回（2）
    import zong
    b=zong.quanXuan()#实现判断订单明细是否能加载成功或者存在
    if not b:return {}
    #zong.xiuGaiBiaoQian()#设置标签为测试
    b=[i.split('\t') for i in b.split('\r\n校验')]#关注前11个--店铺名\仓库名\省\市\区\地址\客服备注\客户备注\快递\异常原因\平台单号
    c={}
    for i in b:
        try:
            if i[12]=='' or i[12]=='测试':
                c[i[11]]={'店铺名':i[1],'仓库名':i[2],'省份':i[3],'城市':i[4],'区':i[5],'地址':i[6],'客服备注':i[7],'客户备注':i[8],'快递':i[9],'异常原因':i[10],'订单标签':i[12]}
        except:print(i,'订单分割存不成功')#;Ri_zhi()
    return c

#加载中必换，其他可以通过控件坐标定位
def ddhpsj():#订单货品数据：获取当前查询到订单的货品数据并返回（3）
    import zong,abc123
    d={}
    f=''
    for j in range(200):
        for i in range(36):#此循环实现判断订单货品明细是否能加载成功
            b=pyperclip.copy("")
            pyautogui.click(*zong.zb['商品信息'])#商品信息
            time.sleep(0.1)
            pyautogui.click(*zong.zb['商品全选'])#全选商品
            abc123.kuaiJieJian(17,67)
            time.sleep(0.1)
            b=pyperclip.paste()
            if re.findall('平台商品|普通商品',b):pan_tiao=0 ;break
        try:
            b=[i.split('\t') for i in b.split('\r\n校验')]#商品编码\品名\数量\规格\商品类型
            b=[{'商品编码':i[1],'品名':i[2],'数量':i[3],'规格':i[4],'商品类型':i[5]} for i in b]#前五个
        except:print('出错87') ; break
        pyautogui.click(*zong.zb['订单信息'])#订单信息
        pyautogui.press('tab',3)
        abc123.kuaiJieJian(17,67)
        if f!=pyperclip.paste():
            f=pyperclip.paste()
        else:break
        if b:d[f]={'订单货品':b}
        else:print(f,'商品不符合条件，或者不存在')
        pyautogui.click(*zong.zb['订单列表'])
        pyautogui.press('down')
    return d

def ddsp():#订单商品：调用ddsj，和ddhpsj 返回订单商品数据
    dd=ddsj()
    print('dd的订单数是：%s'% len(dd))
    if not dd:return []
    data1=ddhpsj()
    if not data1:return []
    print('hp的订单数是：%s'% len(data1))
    hp={}
    for i in data1:
        try:hp[i]=dict(data1[i],**dd[i])
        except:pass#print(i,'订单商品合并失败');Ri_zhi()
    return hp

def readConfigure():
    '''获取配置，返回各配置文件配置'''
    wb=load_workbook('快递配置.xlsx')
    ws=wb['人工处理']
    rgcl=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
    rgcl=[i for i in rgcl if i.count(None)!=len(i)]
    rgcl=[{i1:i2 for i1,i2 in zip(rgcl[0],i) if i2 and i1 !='备注'} for i in rgcl[1:]]
    ws=wb['改快递']
    gkd=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
    gkd=[i for i in gkd if i.count(None)!=len(i)]
    gkd=[{i1:i2 for i1,i2 in zip(gkd[0],i) if i2 and i1 !='备注'} for i in gkd[1:]]
    ws=wb['改赠品']
    gzp=[[i.value for i in ws[i+1]] for i in range(ws.max_row)]
    gzp=[i for i in gzp if i.count(None)!=len(i)]
    gzp=[{i1:i2 for i1,i2 in zip(gzp[0],i) if i2 and i1 !='备注'} for i in gzp[1:]]
    ws=wb['当前使用快递']
    k=[[i.value for i in ws[i+1]] for i in range(ws.max_row)][1:]
    k=[i for i in k if i.count(None)!=len(i)]
    k={i[0]:i[1] for i in k}
    wb.close
    return rgcl,gkd,gzp,k

def panduanpeizhi(v1,v2):#判段配置，传入配置值和订单对应字段值，符合配置返回True,不符合返回False
    #if v1==None:return True
    if re.findall('\d+<%s<\d+',v1):return eval(v1% v2)
    else:
        tj=re.findall('^\[包含\]|^\[不包含\]|^\[空\]|^\[非空\]',v1[:5])[0]
        v1=v1.replace(tj,'')
    if '&' in v1:
        zf01=v1.split('&')
        zf01=['(?=.*%s)'% i for i in zf01]
        v1=''.join(zf01)+'.+'
    if tj=='[包含]':
        if re.findall(v1,v2):return True
        else:return False
    elif tj=='[不包含]':
        if re.findall(v1,v2):return False
        else:return True
    elif tj=='[非空]':
        if v2=='':return False
        else:return True
    elif tj=='[空]':
        if v2=='':return True
        else:return False

def panduan(hp,Config):
    '''传入订单货品数据和配置内容，返回zdsh（自动修改）、zdsh1（自动审核平台单号）、人工处理（平台单号）'''
    rgcl,gkd,gzp,k=Config
    zdsh={'改赠品':[],'关闭订单':[],'改地址':{}}
    zdsh['改快递']={k[i]:[] for i in k}
    zdsh1=[]
    人工处理=[]
    print('将要分析的订单数是：%s'%len(hp))
    for i in hp:#i订单编号
        dianpu001=re.findall('SEPTWOLVES雅赋专卖店|少年狼箱包|拼多多美之瑞专卖店',hp[i]['店铺名'])
        if dianpu001:
            with open("触发日志.txt",'a') as f:f.write('订单%s存在不需要处理的店铺%s\n'%(i,dianpu001))
            continue#这几个店铺不处理
        try:
            guige,pingming,shuliang='','',0
            for j1 in hp[i]['订单货品']:
                if '赠品' in j1['商品类型']:continue
                guige+=j1['规格']
                pingming+=j1['品名']
                shuliang+=int(j1['数量'])
            hp[i]['规格']=guige
            hp[i]['品名']=pingming
            hp[i]['平台商品总数量']=shuliang
            panduan_rg=False
            #筛选改地址
            dz01=re.findall('改地址[:：](.*?)┋',hp[i]['客服备注'])
            if dz01:
                dz01,shengshi=diZhiChuLi(hp[i]['客服备注'])
                if dz01:
                    zdsh['改地址'][i]=dz01#符合条件，修改备注
                    if shengshi:
                        hp[i]['省份']=shengshi[0]
                        hp[i]['城市']=shengshi[1]
                else:
                    with open("触发日志.txt",'a') as f:f.write('订单%s改地址识别原因需要人工审核\n'% i)
                    panduan_rg=True#判断人工处理1

            #判断人工处理2
            r=hp[i]['客服备注'].split('┋')
            kdlx1=''
            kdlx1='|'.join([k1 for k1 in k])
            if len([i7 for i7 in r if  not re.findall(kdlx1+'长度[:：]|改地址[:：]',i7)])>1:
                panduan_rg=True
                with open("触发日志.txt",'a') as f:f.write('订单%s备注含有需要人工处理的内容\n'% i)

            #判断人工处理3
            for i11 in rgcl:#i1一行
                i1={i:i11[i] for i in i11 if not re.findall('处理方式',i)}
                pan1=1
                for i4 in i1:
                    if not panduanpeizhi(i1[i4],hp[i][i4]):#人工筛选
                        pan1=0
                        break
                if pan1:
                    panduan_rg=True
                    with open("触发日志.txt",'a') as f:f.write('订单%s符合条件%s人工处理\n'%(i,i1))
                    break

            #筛选改快递
            for i11 in gkd:#i1一行
                i1={i:i11[i] for i in i11 if not re.findall('使用快递',i)}#当前选择条件（一行）
                pan1=1
                for i4 in i1:
                    if not panduanpeizhi(i1[i4],hp[i][i4]):#快递筛选
                        pan1=0
                        break
                if pan1:
                    if '客服备注指定' not in i11['使用快递']:
                        kd=i11['使用快递']
                    else:kd=re.findall(kdlx1,hp[i]['客服备注'])[-1]
                    if kd not in hp[i]['快递']:
                        zdsh['改快递'][k[kd]].append(i)
                        with open("触发日志.txt",'a') as f:f.write('订单%s符合条件%s改快递为[%s]\n'%(i,i1,kd))
                    break

            #筛选改赠品  
            for i11 in gzp:#i1一行
                i1={i:i11[i] for i in i11 if not re.findall('处理方式|使用快递',i)}
                clfs=i11['处理方式']
                try:
                    sykd01=i11['使用快递']
                except:sykd01=''
                pan1=1
                for i4 in i1:
                    if not panduanpeizhi(i1[i4],hp[i][i4]):
                        pan1=0
                        break
                if pan1:
                    if sykd01=='' or sykd01==kd:
                        zdsh['改赠品'].append(i)
                        with open("触发日志.txt",'a') as f:f.write('订单%s符合条件%s改打火机为卡包\n'%(i,i1))
                        break#向后移动了一格，避免了i1为空的情况
        except:
            panduan_rg=True
            with open("触发日志.txt",'a') as f:f.write('订单%s错误需要人工处理！！\n'% i)
            Ri_zhi()
            try:print(j1,'错误需要人工处理！！',i)
            except:pass
        if panduan_rg:人工处理.append(i)
        else:zdsh1+=[i]
    return zdsh,zdsh1,人工处理#返回需要修改的订单、所有可以自动审核订单、人工处理订单

#******************************************取数据***********************************************

class Ri_zhi():#日志和记录
    def __init__(self):
        now = int(time.time())
        self.timeArray = time.localtime(now)
        self.otherStyleTime = time.strftime("%Y-%m-%d %H:%M:%S",self.timeArray)
        if traceback.format_exc()!='NoneType: None\n':
            f=open(r"执行日志.txt",'a')
            traceback.print_exc(file=f)
            f.write('*******************************'+self.otherStyleTime+'************************************\n\n')
            f.flush()  
            f.close()
    def xujiludnr(self,neirong):
        rq01='%s%s%s'%(self.timeArray[0],self.timeArray[1],self.timeArray[2])
        f=open("错误订单.txt",'a')
        f.write(neirong+'\n')
        f.flush()
        f.close()

if __name__ == '__main__':
    #b,c,d=panduan()
    #b=ddsj()
    pass
    hp=ddsp()
    Config=readConfigure()#获取配置内容
    b,c,d=panduan(hp,Config)


'''
def xgma(i,n,h=1):#西格玛之python实现
    zhi=0
    for j in range(i,n+1):
        zhi+=j*h
    return zhi


'''


